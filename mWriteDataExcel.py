
import soundfile as sf
import numpy as np
import librosa
import random
import time
import os
# file_name = r'C:\Users\3080\Desktop\测试bin10k\tempWav\20151021101527_39_44_中型货车声音_-1.wav'
# X, sample_rate = sf.read(file_name, dtype='float32')
# tt = X.T


from openpyxl import Workbook,load_workbook
import os

class Do_Excel:
    def __init__(self,filename,sheetname='Sheet1'):
        self.filename=filename
        self.sheetname=sheetname

    def write(self,i,j,value):
        if not os.path.exists(self.filename):
            wb = Workbook()
            sh = wb.create_sheet(self.sheetname)
             
        else:
            wb = load_workbook(self.filename)
            sh = wb[self.sheetname]

        if type(value) is np.ndarray:
            [cc,kk] = value.shape[0:2]
            for a in range(cc):
                for b in range(kk):
                    sh.cell(i+a, j+b).value = value[a,b]

        else:
            sh.cell(i,j).value=value


        wb.save(self.filename)

# Do_Excel('test22x2.xlsx').write(1,1,'sdcds')
# Do_Excel('test22x2.xlsx').write(1,2,'change')
# Do_Excel('test22x2.xlsx').write(3,2,'pass')


